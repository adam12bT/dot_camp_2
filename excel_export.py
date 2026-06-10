"""
excel_export.py – Builds the Excel workbook from evaluated startup data.

Sheet layout:
  1. Active Filters     – filter pipeline used for this export
  2. All Startups       – every startup (English + French merged), with Decision column
  3. Age <2 years       – startups in the <2 age group
  4. Age 2–5 years      – startups in the 2-5 age group
  5. Age 5–7 years      – startups in the 5-7 age group
  6. Final Selection    – Selected ★ / Selected only
  7. Charts – Sectors   – sector breakdown table + bar/pie charts
  8. Charts – Overview  – age, decision, maturity, revenue charts

No Streamlit dependency; call build_excel() from anywhere.
"""

import io
from collections import Counter

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint

from engine import evaluate
from utils import bool_str

# ── AGE GROUP normaliser (mirrors engine._AGE_GROUP) ─────────────────────────
_AGE_GROUP = {
    "Less than 2 years":  "<2",
    "Less than  2 years": "<2",
    "Moins de 2 ans":     "<2",
    "2–5 years": "2-5", "2-5 years": "2-5",
    "2–5 ans":   "2-5", "2-5 ans":   "2-5",
    "5–7 years": "5-7", "5-7 years": "5-7",
    "5–7 ans":   "5-7", "5-7 ans":   "5-7",
    "More than 7 years": ">7",
    "Plus de 7 ans":     ">7",
}

def _age_grp(app):
    return _AGE_GROUP.get(str(app.get("age", "")).strip(), "?")


# ── Colour palette ────────────────────────────────────────────────────────────
DARK_NAVY   = "1A1A2E"
ACCENT      = "00E5A0"
LIGHT_GREY  = "F8F9FF"
WHITE       = "FFFFFF"

DECISION_HEX = {
    "Selected ★":    "00E5A0",
    "Selected":      "00B87A",
    "Shortlisted ✓": "F5C842",
    "Shortlisted":   "E5A800",
    "Rejected":      "FF4D6D",
}

AGE_BAND_HEX = {
    "<2":  "D9EAD3",   # soft green
    "2-5": "FCE5CD",   # soft orange
    "5-7": "CFE2F3",   # soft blue
    ">7":  "EAD1DC",   # soft pink
}


# ── Low-level style helpers ───────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, end_color=hex_color)

def _font(bold=False, color=None, size=10, italic=False) -> Font:
    return Font(name="Arial", bold=bold, size=size,
                color=color or "000000", italic=italic)

def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_bottom(color="CCCCCC") -> Border:
    return Border(bottom=Side(style="thin", color=color))

def _write_header_row(ws, row_num: int, headers: list,
                      bg=DARK_NAVY, fg=WHITE, size=10) -> None:
    for c, val in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.fill      = _fill(bg)
        cell.font      = _font(bold=True, color=fg, size=size)
        cell.alignment = _align(h="center")
        cell.border    = _border_bottom()
    ws.row_dimensions[row_num].height = 28


def _zebra_row(ws, row_num: int, values: list, even: bool,
               height=18, decision_col: int = None, decision_val: str = None) -> None:
    bg = LIGHT_GREY if even else WHITE
    for c, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=c, value=val)
        cell.fill      = _fill(bg)
        cell.font      = _font(size=9)
        cell.alignment = _align(v="center")
    # Colour the decision cell if provided
    if decision_col and decision_val:
        dc = ws.cell(row=row_num, column=decision_col)
        hex_c = DECISION_HEX.get(decision_val, "CCCCCC")
        dc.fill = _fill(hex_c)
        is_light = decision_val in ("Selected ★", "Selected", "Shortlisted ✓", "Shortlisted")
        dc.font = _font(bold=True, size=9, color="001A10" if is_light else WHITE)
    ws.row_dimensions[row_num].height = height


def _freeze_and_filter(ws, headers_row=1, last_col_letter=None) -> None:
    ws.freeze_panes = f"A{headers_row + 1}"
    if last_col_letter:
        ws.auto_filter.ref = f"A{headers_row}:{last_col_letter}{headers_row}"


def _section_title(ws, row_num: int, title: str, span: int = 4) -> None:
    ws.cell(row=row_num, column=1, value=title).font = _font(bold=True, size=13)
    ws.merge_cells(start_row=row_num, start_column=1,
                   end_row=row_num, end_column=span)
    ws.row_dimensions[row_num].height = 22


# ── Column schema for startup rows ───────────────────────────────────────────

STARTUP_COLS = [
    ("Startup Name",       lambda a, r: a.get("startup_name", "")),
    ("Age",                lambda a, r: a.get("age", "")),
    ("Age Group",          lambda a, r: _age_grp(a)),
    ("Sector",             lambda a, r: a.get("sector", "")),
    ("Governorate",        lambda a, r: a.get("governorate", "")),
    ("Maturity",           lambda a, r: a.get("maturity", "")),
    ("Employees",          lambda a, r: a.get("total_employees", "")),
    ("Clients",            lambda a, r: a.get("num_clients", "")),
    ("Full-time Founder",  lambda a, r: bool_str(a.get("full_time_founder", False))),
    ("Founder in TN",      lambda a, r: bool_str(a.get("founder_in_tunisia", False))),
    ("Gender Mixed",       lambda a, r: bool_str(a.get("gender_mixed", False))),
    ("Legally Created",    lambda a, r: bool_str(a.get("legally_created", False))),
    ("Revenue",            lambda a, r: bool_str(a.get("generating_revenue", False))),
    ("Funding",            lambda a, r: bool_str(a.get("raised_funding", False))),
    ("Startup Label",      lambda a, r: bool_str(a.get("startup_label", False))),
    ("Bonus Score",        lambda a, r: r.bonus_score),
    ("Decision",           lambda a, r: r.final_decision),
    ("Rejection Reason",   lambda a, r: r.rejection_reason or ""),
]

COL_WIDTHS_STARTUP = {
    "A": 28, "B": 18, "C": 12, "D": 16, "E": 14,
    "F": 26, "G": 12, "H": 12, "I": 16, "J": 14,
    "K": 14, "L": 15, "M": 10, "N": 10, "O": 14,
    "P": 12, "Q": 16, "R": 36,
}

def _startup_headers(filter_labels: list[str]) -> list[str]:
    base = [h for h, _ in STARTUP_COLS[:-2]]   # everything before Decision
    filter_hdrs = [f"Filter: {l}" for l in filter_labels]
    return base + filter_hdrs + ["Decision", "Rejection Reason"]

def _startup_row_values(app: dict, result, filter_labels: list[str]) -> tuple[list, int, str]:
    """Returns (values_list, decision_col_index_1based, decision_value)."""
    base_vals = [fn(app, result) for _, fn in STARTUP_COLS[:-2]]
    fr_map    = {lbl: res for lbl, res in result.filter_results}
    filter_vals = [fr_map.get(l, "—") for l in filter_labels]
    dec_col = len(base_vals) + len(filter_labels) + 1   # 1-based
    return base_vals + filter_vals + [result.final_decision, result.rejection_reason or ""], dec_col, result.final_decision


def _set_startup_col_widths(ws, n_base: int, n_filters: int) -> None:
    for letter, w in COL_WIDTHS_STARTUP.items():
        ws.column_dimensions[letter].width = w
    # Filter columns
    for i in range(n_base + 1, n_base + n_filters + 1):
        ws.column_dimensions[get_column_letter(i)].width = 30
    # Tail
    tail = n_base + n_filters
    ws.column_dimensions[get_column_letter(tail + 1)].width = 16   # Decision
    ws.column_dimensions[get_column_letter(tail + 2)].width = 40   # Rejection reason


def _write_startup_sheet(ws, rows: list[tuple], filter_labels: list[str],
                          age_highlight: str = None) -> None:
    """Write a full startup table into ws. rows = [(app, result), ...]"""
    headers  = _startup_headers(filter_labels)
    n_base   = len(STARTUP_COLS) - 2        # cols before filter cols
    n_filter = len(filter_labels)

    _write_header_row(ws, 1, headers)
    _freeze_and_filter(ws, 1, get_column_letter(len(headers)))
    _set_startup_col_widths(ws, n_base, n_filter)

    for r_idx, (app, result) in enumerate(rows, 2):
        vals, dec_col, dec_val = _startup_row_values(app, result, filter_labels)

        # Optional: tint age-group column if filtering by one group
        even = r_idx % 2 == 0
        _zebra_row(ws, r_idx, vals, even,
                   decision_col=dec_col, decision_val=dec_val)

        # Highlight age-group cell (col C = index 3)
        if age_highlight:
            ag_cell = ws.cell(row=r_idx, column=3)
            ag_hex  = AGE_BAND_HEX.get(age_highlight, LIGHT_GREY)
            ag_cell.fill = _fill(ag_hex)


# ── Sheet 1 – Active Filters ──────────────────────────────────────────────────

def _build_filters_sheet(ws, active_config: dict) -> None:
    import json

    def _make_serializable(obj):
        if isinstance(obj, dict):
            return {str(k) if isinstance(k, tuple) else k: _make_serializable(v)
                    for k, v in obj.items()}
        return obj

    _section_title(ws, 1, f"Active Filter Pipeline  ·  Config: {active_config.get('name','')}", span=5)

    headers = ["#", "Label", "Type", "Params (JSON)", "Description"]
    _write_header_row(ws, 3, headers)

    TYPE_DESC = {
        "age_reject":       "Reject startups in specified age groups",
        "legal_check":      "Require legal registration for specified age groups",
        "boolean_field":    "Reject if a single boolean field is False",
        "multi_boolean":    "Reject unless all/any of several boolean fields are True",
        "maturity_outcome": "Map (age group × maturity) to Selected / Shortlisted",
        "field_in":         "Reject if field value is not in allowed set",
        "field_not_in":     "Reject if field value is in blocked set",
        "min_employees":    "Reject if employee band is below minimum",
    }

    for i, f in enumerate(active_config.get("filters", []), 4):
        even = i % 2 == 0
        params_str = json.dumps(_make_serializable(f.get("params", {})), default=str)
        vals = [
            i - 3,
            f.get("label", ""),
            f.get("type", ""),
            params_str,
            TYPE_DESC.get(f.get("type", ""), ""),
        ]
        _zebra_row(ws, i, vals, even, height=20)

    # Bonus config block
    gap = len(active_config.get("filters", [])) + 6
    _section_title(ws, gap, "Bonus Scoring Config", span=3)
    _write_header_row(ws, gap + 1, ["Field", "Label", "Function", "Active"])
    for j, (fk, lbl, fn) in enumerate(active_config.get("bonus_fields", []), gap + 2):
        _zebra_row(ws, j, [fk, lbl, fn, "Yes"], j % 2 == 0, height=18)

    gap2 = gap + 2 + len(active_config.get("bonus_fields", []))
    ws.cell(row=gap2 + 1, column=1, value=f"★ threshold:  {active_config.get('star_threshold', 3)}").font = _font(bold=True)
    ws.cell(row=gap2 + 2, column=1, value=f"✓ threshold:  {active_config.get('checkmark_threshold', 2)}").font = _font(bold=True)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 45


# ── Sheet 7 & 8 – Charts ─────────────────────────────────────────────────────

def _build_sector_chart_sheet(ws, evaluated: list) -> None:
    """Sector breakdown table + bar chart + pie chart."""
    sectors = Counter(a.get("sector", "Unknown") or "Unknown" for a, _ in evaluated)
    sorted_sectors = sorted(sectors.items(), key=lambda x: -x[1])

    _section_title(ws, 1, "Sector Breakdown", span=2)
    _write_header_row(ws, 2, ["Sector", "Count"])

    for i, (sec, cnt) in enumerate(sorted_sectors, 3):
        ws.cell(row=i, column=1, value=sec).font  = _font(size=10)
        ws.cell(row=i, column=2, value=cnt).font  = _font(size=10)
        ws.row_dimensions[i].height = 18

    total_row = 3 + len(sorted_sectors)
    ws.cell(row=total_row, column=1, value="TOTAL").font = _font(bold=True, size=10)
    ws.cell(row=total_row, column=2, value=sum(sectors.values())).font = _font(bold=True, size=10)

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 10

    last_data_row = total_row - 1

    # Bar chart
    bar = BarChart()
    bar.type       = "col"
    bar.title      = "Startups by Sector"
    bar.y_axis.title = "Count"
    bar.x_axis.title = "Sector"
    bar.style      = 10
    bar.width      = 22
    bar.height     = 14
    data_ref  = Reference(ws, min_col=2, min_row=2, max_row=last_data_row)
    cats_ref  = Reference(ws, min_col=1, min_row=3, max_row=last_data_row)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.series[0].title = None
    ws.add_chart(bar, "D2")

    # Pie chart
    pie = PieChart()
    pie.title  = "Sector Distribution"
    pie.style  = 10
    pie.width  = 18
    pie.height = 14
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)
    ws.add_chart(pie, "D20")


def _build_overview_chart_sheet(ws, evaluated: list) -> None:
    """Age groups + Decision + Maturity + Revenue/Funding charts."""

    # ── Age group counts ──────────────────────────────────────────────────────
    AGE_LABELS = {"<2": "< 2 years", "2-5": "2–5 years", "5-7": "5–7 years", ">7": "> 7 years"}
    age_counts = Counter(_age_grp(a) for a, _ in evaluated)

    _section_title(ws, 1, "Age Group Breakdown", span=2)
    _write_header_row(ws, 2, ["Age Group", "Count"])
    age_order = ["<2", "2-5", "5-7", ">7"]
    for i, ag in enumerate(age_order, 3):
        ws.cell(row=i, column=1, value=AGE_LABELS[ag]).font = _font(size=10)
        ws.cell(row=i, column=2, value=age_counts.get(ag, 0)).font = _font(size=10)
        ws.row_dimensions[i].height = 18

    bar_age = BarChart()
    bar_age.type   = "col"
    bar_age.title  = "Startups by Age Group"
    bar_age.style  = 10
    bar_age.width  = 18
    bar_age.height = 12
    bar_age.add_data(Reference(ws, min_col=2, min_row=2, max_row=6), titles_from_data=True)
    bar_age.set_categories(Reference(ws, min_col=1, min_row=3, max_row=6))
    ws.add_chart(bar_age, "D1")

    # ── Decision counts ───────────────────────────────────────────────────────
    dec_order  = ["Selected ★", "Selected", "Shortlisted ✓", "Shortlisted", "Rejected"]
    dec_counts = Counter(r.final_decision for _, r in evaluated)

    row_offset = 9
    _section_title(ws, row_offset, "Decision Breakdown", span=2)
    _write_header_row(ws, row_offset + 1, ["Decision", "Count"])
    for i, d in enumerate(dec_order, row_offset + 2):
        ws.cell(row=i, column=1, value=d).font  = _font(size=10)
        ws.cell(row=i, column=2, value=dec_counts.get(d, 0)).font = _font(size=10)
        ws.row_dimensions[i].height = 18

    pie_dec = PieChart()
    pie_dec.title  = "Decision Distribution"
    pie_dec.style  = 10
    pie_dec.width  = 18
    pie_dec.height = 12
    pie_dec.add_data(Reference(ws, min_col=2, min_row=row_offset + 1,
                               max_row=row_offset + 1 + len(dec_order)), titles_from_data=True)
    pie_dec.set_categories(Reference(ws, min_col=1, min_row=row_offset + 2,
                                     max_row=row_offset + 1 + len(dec_order)))
    ws.add_chart(pie_dec, "D18")

    # ── Maturity counts ───────────────────────────────────────────────────────
    from engine import MATURITY_OPTIONS
    mat_counts = Counter(a.get("maturity", "Unknown") for a, _ in evaluated)

    row_m = row_offset + len(dec_order) + 4
    _section_title(ws, row_m, "Maturity Breakdown", span=2)
    _write_header_row(ws, row_m + 1, ["Maturity Stage", "Count"])
    for i, m in enumerate(MATURITY_OPTIONS, row_m + 2):
        ws.cell(row=i, column=1, value=m).font  = _font(size=10)
        ws.cell(row=i, column=2, value=mat_counts.get(m, 0)).font = _font(size=10)
        ws.row_dimensions[i].height = 18

    bar_mat = BarChart()
    bar_mat.type   = "bar"   # horizontal
    bar_mat.title  = "Startups by Maturity"
    bar_mat.style  = 10
    bar_mat.width  = 20
    bar_mat.height = 14
    bar_mat.add_data(Reference(ws, min_col=2, min_row=row_m + 1,
                               max_row=row_m + 1 + len(MATURITY_OPTIONS)), titles_from_data=True)
    bar_mat.set_categories(Reference(ws, min_col=1, min_row=row_m + 2,
                                     max_row=row_m + 1 + len(MATURITY_OPTIONS)))
    ws.add_chart(bar_mat, "D35")

    # ── Revenue / Funding ─────────────────────────────────────────────────────
    rev_yes = sum(1 for a, _ in evaluated if a.get("generating_revenue"))
    rev_no  = len(evaluated) - rev_yes
    fund_yes = sum(1 for a, _ in evaluated if a.get("raised_funding"))
    fund_no  = len(evaluated) - fund_yes

    row_rf = row_m + len(MATURITY_OPTIONS) + 4
    _section_title(ws, row_rf, "Revenue & Funding", span=2)
    _write_header_row(ws, row_rf + 1, ["Category", "Yes", "No"])
    ws.cell(row=row_rf + 2, column=1, value="Generating Revenue").font = _font(size=10)
    ws.cell(row=row_rf + 2, column=2, value=rev_yes).font  = _font(size=10)
    ws.cell(row=row_rf + 2, column=3, value=rev_no).font   = _font(size=10)
    ws.cell(row=row_rf + 3, column=1, value="Raised Funding").font    = _font(size=10)
    ws.cell(row=row_rf + 3, column=2, value=fund_yes).font = _font(size=10)
    ws.cell(row=row_rf + 3, column=3, value=fund_no).font  = _font(size=10)

    bar_rf = BarChart()
    bar_rf.type   = "col"
    bar_rf.title  = "Revenue & Funding"
    bar_rf.style  = 10
    bar_rf.width  = 18
    bar_rf.height = 12
    bar_rf.add_data(Reference(ws, min_col=2, min_row=row_rf + 1, max_col=3, max_row=row_rf + 3),
                    titles_from_data=True)
    bar_rf.set_categories(Reference(ws, min_col=1, min_row=row_rf + 2, max_row=row_rf + 3))
    ws.add_chart(bar_rf, "D53")

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10


# ── Public entry-point ────────────────────────────────────────────────────────

def build_excel(startups: list, active_config: dict) -> bytes:
    """
    Generate Excel workbook with 8 sheets:
      1. Active Filters
      2. All Startups
      3. Age < 2 years
      4. Age 2–5 years
      5. Age 5–7 years
      6. Final Selection
      7. Charts – Sectors
      8. Charts – Overview
    """
    wb = Workbook()
    wb.remove(wb.active)

    filter_labels = [f["label"] for f in active_config.get("filters", [])]
    evaluated     = [(app, evaluate(app, config=active_config)) for app in startups]

    # 1. Active Filters
    ws_flt = wb.create_sheet("Active Filters")
    _build_filters_sheet(ws_flt, active_config)

    # 2. All Startups (English + French merged)
    ws_all = wb.create_sheet("All Startups")
    _write_startup_sheet(ws_all, evaluated, filter_labels)

    # 3–5. Age-group sheets
    age_sheet_defs = [
        ("<2",  "Age  <2 years"),
        ("2-5", "Age 2–5 years"),
        ("5-7", "Age 5–7 years"),
    ]
    for ag_code, sheet_name in age_sheet_defs:
        subset = [(a, r) for a, r in evaluated if _age_grp(a) == ag_code]
        ws_ag  = wb.create_sheet(sheet_name)
        _write_startup_sheet(ws_ag, subset, filter_labels, age_highlight=ag_code)

    # 6. Final Selection (Selected ★ + Selected)
    selected = [(a, r) for a, r in evaluated
                if r.final_decision in ("Selected ★", "Selected")]
    ws_sel = wb.create_sheet("Final Selection")
    _write_startup_sheet(ws_sel, selected, filter_labels)

    # 7. Charts – Sectors
    ws_sec = wb.create_sheet("Charts – Sectors")
    _build_sector_chart_sheet(ws_sec, evaluated)

    # 8. Charts – Overview
    ws_ov = wb.create_sheet("Charts – Overview")
    _build_overview_chart_sheet(ws_ov, evaluated)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
