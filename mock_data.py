"""
mock_data.py – Data generators + importers for Dot Camp selection pipeline.

Supports:
  - Random batch generation
  - Typeform CSV export (English columns)
  - Typeform CSV export (French columns)
  - Typeform JSON export
  - Google Form CSV export (legacy)
"""

import random
import json
import csv
import io
from datetime import datetime, timedelta

from engine import AGE_OPTIONS, MATURITY_OPTIONS, EMPLOYEE_RANGES, CLIENT_RANGES

random.seed(42)

SECTORS = [
    "AI/Data", "FinTech", "HealthTech", "EdTech", "GreenTech",
    "Agri/FoodTech", "HR Tech", "DeepTech", "SaaS", "TravelTech",
    "PropTech", "Industry 4.0", "ICC", "Transport Tech",
]

GOVERNORATES = [
    "Tunis", "Ariana", "Ben Arous", "Manouba", "Nabeul",
    "Bizerte", "Sousse", "Monastir", "Sfax", "Kairouan",
    "Mahdia", "Médenine", "Gafsa", "Jendouba",
]

STARTUP_NAMES = [
    "NovaSpark", "LumiPay", "TerraFarm", "MediPulse", "UrbanFlow",
    "ClearDoc", "SmartCraft", "EduBoost", "CarbonZero", "AgroMind",
    "HealthBridge", "TalentCore", "DataLens", "PayEase", "GreenVault",
    "CodeNest", "SolarLink", "MarketEye", "SkillUp", "BioTrack",
    "CloudFactory", "FoodRescue", "LearnPath", "EnergyPulse", "RoboFarm",
    "QuickCare", "HireAI", "CropSense", "FinGuru", "SafeRoute",
]

# ── Bool normalisation ────────────────────────────────────────────────────────
_BOOL_YES = {"yes", "oui", "true", "1", "1.0", "vrai", "o", "y"}

def _to_bool(v) -> bool:
    if isinstance(v, bool): return v
    return str(v).strip().lower() in _BOOL_YES

def _norm_range(v: str) -> str:
    """Normalise employee/client range strings to engine format."""
    s = str(v).strip() if v else ""
    if not s or s.lower() in {"none", "aucun", "0", "nan", "null", ""}:
        return "None"
    # Handle numeric strings
    for band in ["1–2", "3–5", "6–10", "+10"]:
        if band in s or band.replace("–", "-") in s:
            return band
    # Typeform sometimes returns plain numbers
    try:
        n = int(float(s.replace("+", "").replace(" ", "")))
        if n == 0:             return "None"
        if 1 <= n <= 2:        return "1–2"
        if 3 <= n <= 5:        return "3–5"
        if 6 <= n <= 10:       return "6–10"
        if n > 10:             return "+10"
    except ValueError:
        pass
    return "None"

def _norm_age(v: str) -> str:
    """Normalise age strings to engine AGE_OPTIONS."""
    s = str(v).strip().lower() if v else ""
    if "moins" in s or "less" in s or "<2" in s or "2 ans" not in s and any(x in s for x in ["1 an", "moins", "less", "2 year"]):
        return "Less than 2 years"
    if "2" in s and ("5" in s or "cinq" in s):
        return "2–5 years"
    if "5" in s and ("7" in s or "sept" in s):
        return "5–7 years"
    if "plus" in s or "more" in s or ">7" in s or "7" in s:
        return "More than 7 years"
    # Direct match
    mapping = {
        "less than 2 years":  "Less than 2 years",
        "less than  2 years": "Less than 2 years",
        "moins de 2 ans":     "Less than 2 years",
        "2–5 years": "2–5 years", "2-5 years": "2–5 years",
        "2–5 ans":   "2–5 years", "2-5 ans":   "2–5 years",
        "5–7 years": "5–7 years", "5-7 years": "5–7 years",
        "5–7 ans":   "5–7 years", "5-7 ans":   "5–7 years",
        "more than 7 years":  "More than 7 years",
        "plus de 7 ans":      "More than 7 years",
    }
    return mapping.get(s, v)  # fall back to raw value if unrecognised

def _norm_maturity(v: str) -> str:
    s = str(v).strip() if v else ""
    mapping = {
        "Idée":                                          "Idea",
        "Stade d'idée":                                  "Idea",
        "Stade d\u2019id\u00e9e\xa0":                   "Idea",
        "POC finalisé":                                  "POC finalized",
        "Poc finalisé":                                  "POC finalized",
        "MVP fonctionnel":                               "Functional MVP",
        "MVP en cours de test":                          "MVP currently being tested",
        "Go To Market (Premières ventes)":               "Go To Market (Early sales)",
        "Go To Market (Eraly salees)":                   "Go To Market (Early sales)",
        "Produit / Service sur le marché":               "Product/Service on the market",
        "Produit/ Service sur le marché":                "Product/Service on the market",
        "Produit / Service commercialisé":               "Product/Service on the market",
        "Produit / Service en expansion internationale": "International Expansion",
        "Produit/ Service en expansion internationale":  "International Expansion",
    }
    return mapping.get(s, s)


# ── Column index maps (0-based) for Typeform CSV ──────────────────────────────
# Both EN and FR sheets share the same column positions — only the header text differs.
# Positions determined from the real Excel export.

_COL = {
    "startup_name":          2,
    "age":                   3,
    "governorate":           4,
    "sector":                5,
    "sector_other":          6,
    "legally_created":       7,
    "founded_year":          8,
    "website":              11,
    "linkedin":             12,
    "full_time_founder":    38,
    "founder_in_tunisia":   39,
    "total_employees":      40,
    "salaried_employees":   41,
    "gender_mixed":         42,
    "problem":              43,
    "solution":             44,
    "maturity":             48,
    "num_clients":          49,
    "generating_revenue":   50,
    "raised_funding":       56,
    "startup_label":        66,
    "participated_programs":70,
    "submitted_at":         95,
}

_BOOL_COLS = {
    "legally_created", "full_time_founder", "founder_in_tunisia",
    "gender_mixed", "generating_revenue", "raised_funding",
    "startup_label", "participated_programs",
}

_RANGE_COLS = {"total_employees", "salaried_employees", "num_clients"}


def _row_to_app(row: list) -> dict | None:
    """Convert a raw CSV/xlsx row (list of values) into an engine-ready dict."""
    def get(field):
        idx = _COL.get(field)
        if idx is None or idx >= len(row):
            return ""
        v = row[idx]
        return "" if v is None else str(v).strip()

    name = get("startup_name")
    if not name:
        return None

    app = {"startup_name": name}

    for field in _COL:
        if field == "startup_name":
            continue
        raw = get(field)
        if field in _BOOL_COLS:
            app[field] = _to_bool(raw)
        elif field in _RANGE_COLS:
            app[field] = _norm_range(raw)
        elif field == "age":
            app[field] = _norm_age(raw)
        elif field == "maturity":
            app[field] = _norm_maturity(raw)
        else:
            app[field] = raw

    # Sector: prefer "Other" text if sector col is empty
    if not app.get("sector") or app["sector"].lower() in {"other", "autre", ""}:
        other = get("sector_other")
        if other:
            app["sector"] = other

    return app


# ── Typeform CSV (the real export format) ────────────────────────────────────

def typeform_csv_to_apps(csv_text: str) -> list[dict]:
    """
    Parse a Typeform CSV export (English or French) into engine-ready dicts.
    Handles both languages — column positions are identical.
    Skips the header row automatically.
    """
    reader = csv.reader(io.StringIO(csv_text))
    rows   = list(reader)
    if not rows:
        return []

    apps = []
    # Row 0 is headers — skip it
    for raw_row in rows[1:]:
        app = _row_to_app(raw_row)
        if app:
            apps.append(app)
    return apps


# ── Legacy Google Form CSV ────────────────────────────────────────────────────

GFORM_COLUMN_MAP = {
    "Timestamp":                                          "submitted_at",
    "What is the name of your startup?":                  "startup_name",
    "Age of your startup":                                "age",
    "Sector of your startup":                             "sector",
    "Governorate":                                        "governorate",
    "Is your startup legally established?":               "legally_created",
    "Is one of the founders working full-time?":          "full_time_founder",
    "Is at least one founder a resident of Tunisia?":     "founder_in_tunisia",
    "Total number of employees":                          "total_employees",
    "Number of salaried employees":                       "salaried_employees",
    "Is your team gender-mixed?":                         "gender_mixed",
    "How advanced is your solution?":                     "maturity",
    "Current number of clients/users":                    "num_clients",
    "Is your startup currently generating revenue?":      "generating_revenue",
    "Does your startup have a startup label?":            "startup_label",
    "Has your startup raised funding?":                   "raised_funding",
    "Participated in acceleration programs?":             "participated_programs",
}

_GFORM_BOOL_FIELDS = {
    "legally_created", "full_time_founder", "founder_in_tunisia",
    "gender_mixed", "generating_revenue", "startup_label",
    "raised_funding", "participated_programs",
}

def gform_csv_to_apps(csv_text: str) -> list[dict]:
    """Parse a legacy Google-Form CSV into engine-ready dicts."""
    reader = csv.DictReader(io.StringIO(csv_text))
    apps   = []
    for row in reader:
        app = {}
        for col, field in GFORM_COLUMN_MAP.items():
            val = row.get(col, "")
            if field in _GFORM_BOOL_FIELDS:
                app[field] = _to_bool(val)
            elif field in _RANGE_COLS:
                app[field] = _norm_range(val)
            elif field == "age":
                app[field] = _norm_age(val)
            elif field == "maturity":
                app[field] = _norm_maturity(val)
            else:
                app[field] = val
        if app.get("startup_name"):
            apps.append(app)
    return apps


# ── Typeform JSON ─────────────────────────────────────────────────────────────

TYPEFORM_FIELD_IDS = {
    "startup_name":          "field_001",
    "age":                   "field_002",
    "sector":                "field_003",
    "governorate":           "field_004",
    "legally_created":       "field_005",
    "full_time_founder":     "field_006",
    "founder_in_tunisia":    "field_007",
    "total_employees":       "field_008",
    "salaried_employees":    "field_009",
    "gender_mixed":          "field_010",
    "maturity":              "field_011",
    "num_clients":           "field_012",
    "generating_revenue":    "field_013",
    "startup_label":         "field_014",
    "raised_funding":        "field_015",
    "participated_programs": "field_016",
}

def typeform_json_to_apps(json_text: str) -> list[dict]:
    """Parse Typeform JSON export into engine-ready dicts."""
    data        = json.loads(json_text)
    bool_fields = {
        "legally_created", "full_time_founder", "founder_in_tunisia",
        "gender_mixed", "generating_revenue", "startup_label",
        "raised_funding", "participated_programs",
    }
    apps = []
    for item in data.get("items", []):
        app = {"submitted_at": item.get("submitted_at", "")}
        for ans in item.get("answers", []):
            ref = ans.get("field", {}).get("ref", "")
            if not ref:
                continue
            if ref in bool_fields:
                app[ref] = ans.get("boolean", False)
            else:
                app[ref] = ans.get("text", "")
        if app.get("startup_name"):
            apps.append(app)
    return apps


# ── Serialisers (for testing / mock CSV generation) ──────────────────────────

def apps_to_gform_csv(apps: list[dict]) -> str:
    headers = list(GFORM_COLUMN_MAP.keys())
    inv_map = {v: k for k, v in GFORM_COLUMN_MAP.items()}
    out     = io.StringIO()
    writer  = csv.DictWriter(out, fieldnames=headers)
    writer.writeheader()
    for app in apps:
        row = {}
        for col, field in GFORM_COLUMN_MAP.items():
            val = app.get(field, "")
            if isinstance(val, bool):
                val = "Yes" if val else "No"
            row[col] = val
        writer.writerow(row)
    return out.getvalue()


def apps_to_typeform_json(apps: list[dict]) -> str:
    responses = []
    for i, app in enumerate(apps):
        answers = []
        for field, field_id in TYPEFORM_FIELD_IDS.items():
            val = app.get(field, "")
            if isinstance(val, bool):
                answers.append({"field": {"id": field_id, "ref": field}, "type": "boolean", "boolean": val})
            else:
                answers.append({"field": {"id": field_id, "ref": field}, "type": "text", "text": str(val)})
        responses.append({
            "response_id":  f"resp_{i:04d}",
            "submitted_at": app.get("submitted_at", ""),
            "answers":      answers,
        })
    return json.dumps({"total_items": len(responses), "items": responses}, indent=2)


# ── Random generator ──────────────────────────────────────────────────────────

def _rand_bool(prob_true: float = 0.6) -> bool:
    return random.random() < prob_true

def _rand_range(options, weights=None) -> str:
    return random.choices(options, weights=weights, k=1)[0]

def generate_one(name: str = None, seed_override: int = None) -> dict:
    if seed_override is not None:
        random.seed(seed_override)

    age = _rand_range(AGE_OPTIONS, weights=[35, 35, 15, 5])

    legal_prob = {
        "Less than 2 years": 0.70, "2–5 years": 0.90,
        "5–7 years": 0.95, "More than 7 years": 0.98,
    }
    legally_created = _rand_bool(legal_prob.get(age, 0.8))

    maturity_weights = {
        "Less than 2 years": [10, 20, 30, 25, 10, 4, 1],
        "2–5 years":         [2,  8,  20, 30, 25, 12, 3],
        "5–7 years":         [1,  3,  10, 20, 30, 25, 11],
        "More than 7 years": [0,  1,   5, 10, 25, 40, 19],
    }
    maturity  = _rand_range(MATURITY_OPTIONS, weights=maturity_weights.get(age))
    mat_score = MATURITY_OPTIONS.index(maturity)
    rev_prob  = 0.05 + mat_score * 0.12
    client_w  = [max(1, 10 - mat_score * 2)] + [4, 4 + mat_score, 2 + mat_score, 1 + mat_score]
    emp_w     = [max(1, 8 - mat_score)] + [5, 3 + mat_score, 1 + mat_score // 2, 1]

    return {
        "startup_name":          name or random.choice(STARTUP_NAMES) + str(random.randint(1, 99)),
        "age":                   age,
        "sector":                random.choice(SECTORS),
        "governorate":           random.choice(GOVERNORATES),
        "legally_created":       legally_created,
        "full_time_founder":     _rand_bool(0.78),
        "founder_in_tunisia":    _rand_bool(0.82),
        "total_employees":       _rand_range(EMPLOYEE_RANGES, weights=emp_w),
        "salaried_employees":    _rand_range(EMPLOYEE_RANGES[:-1], weights=[4, 4, 3, 2]),
        "gender_mixed":          _rand_bool(0.52),
        "maturity":              maturity,
        "num_clients":           _rand_range(CLIENT_RANGES, weights=client_w),
        "generating_revenue":    _rand_bool(rev_prob),
        "startup_label":         _rand_bool(0.15),
        "raised_funding":        _rand_bool(0.18),
        "participated_programs": _rand_bool(0.22),
        "submitted_at":          (datetime.now() - timedelta(days=random.randint(0, 30))).strftime("%Y-%m-%d %H:%M"),
    }

def generate_batch(n: int = 30) -> list[dict]:
    random.seed(0)
    return [generate_one() for _ in range(n)]


# ── Smoke test ────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    apps = generate_batch(5)

    # GForm CSV round-trip
    csv_str = apps_to_gform_csv(apps)
    parsed  = gform_csv_to_apps(csv_str)
    print(f"GForm CSV round-trip: {len(parsed)} apps ✓")

    # Typeform JSON round-trip
    tf_str  = apps_to_typeform_json(apps)
    parsed2 = typeform_json_to_apps(tf_str)
    print(f"Typeform JSON round-trip: {len(parsed2)} apps ✓")

    # Typeform CSV from real Excel
    try:
        from openpyxl import load_workbook
        import csv as _csv, io as _io
        wb = load_workbook("/mnt/user-data/uploads/Selection_des_startups_Dot_Camp_5__2_.xlsx")
        ws = wb["All Anglais"]
        buf = _io.StringIO()
        writer = _csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([v if v is not None else "" for v in row])
        real_apps = typeform_csv_to_apps(buf.getvalue())
        print(f"Real Typeform CSV (English sheet): {len(real_apps)} apps parsed ✓")
        if real_apps:
            a = real_apps[0]
            print(f"  First app: {a['startup_name']} | age={a['age']} | maturity={a['maturity']} | revenue={a['generating_revenue']}")
    except Exception as e:
        print(f"Real CSV test skipped: {e}")
